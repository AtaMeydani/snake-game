import openpyxl
from openpyxl.chart import LineChart, Reference, Series


class Dataset:
    NAME = 'a_star_vs_BFS.xlsx'

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.sheet.title = "A-star vs BFS algorithm"
        self.a_star_time_cell = self.sheet.cell(row=1, column=1, value='a_star_time')

        self.a_star_expanded_cubes_cell = self.sheet.cell(row=1, column=2, value='a_star_expanded_cubes')
        self.bfs_time_cell = self.sheet.cell(row=1, column=3, value='bfs_time')
        self.bfs_expanded_cubes_cell = self.sheet.cell(row=1, column=4, value='bfs_expanded_cubes')
        self.wb.save(Dataset.NAME)

    def add(self, a_star_time, a_star_expanded_cubes, bfs_time, bfs_expanded_cubes):
        last_row = self.sheet.max_row + 1
        self.sheet.cell(row=last_row, column=self.a_star_time_cell.column, value=a_star_time)
        self.sheet.cell(row=last_row, column=self.a_star_expanded_cubes_cell.column, value=a_star_expanded_cubes)
        self.sheet.cell(row=last_row, column=self.bfs_time_cell.column, value=bfs_time)
        self.sheet.cell(row=last_row, column=self.bfs_expanded_cubes_cell.column, value=bfs_expanded_cubes)
        self.wb.save(Dataset.NAME)

    def plot(self):
        line_chart_1 = LineChart()
        a_star_time = Reference(self.sheet, min_col=1, min_row=2, max_row=self.sheet.max_row)
        bfs_time = Reference(self.sheet, min_col=3, min_row=2, max_row=self.sheet.max_row)
        a_star_time_series = Series(values=a_star_time, title="a_star_time")
        bfs_time_series = Series(values=bfs_time, title="bfs_time")

        line_chart_1.series.append(a_star_time_series)
        line_chart_1.series.append(bfs_time_series)

        line_chart_1.title = "Comparing Elapsed Time"
        line_chart_1.x_axis.title = "Frequency of algorithm execution"
        line_chart_1.y_axis.title = "Time"

        line_chart_2 = LineChart()
        a_star_expanded_cubes = Reference(self.sheet, min_col=2, min_row=2, max_row=self.sheet.max_row)
        bfs_expanded_cubes = Reference(self.sheet, min_col=4, min_row=2, max_row=self.sheet.max_row)
        a_star_expanded_cubes_series = Series(values=a_star_expanded_cubes, title="a_star_expanded_cubes")
        bfs_expanded_cubes_series = Series(values=bfs_expanded_cubes, title="bfs_expanded_cubes")

        line_chart_2.series.append(a_star_expanded_cubes_series)
        line_chart_2.series.append(bfs_expanded_cubes_series)

        line_chart_2.title = "Comparing Number Of Expanded Cubes"
        line_chart_2.x_axis.title = "Frequency of algorithm execution"
        line_chart_2.y_axis.title = "Cubes"

        self.sheet.add_chart(line_chart_1, "F2")
        self.sheet.add_chart(line_chart_2, "F20")

        self.wb.save(Dataset.NAME)
